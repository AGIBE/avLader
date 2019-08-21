# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, unicode_literals
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Side, Border, Alignment

wb = Workbook()

#Fonts definieren
def style_range(ws, cell_range, fill=None, font=None, alignment=None, border=None):
	rows = ws[cell_range]
	if font:
		for row in rows:
			for c in row:
				c.font = font
	if fill:
		for row in rows:
			for c in row:
				c.fill = fill
	if alignment:
		for row in rows:
			for c in row:
				c.alignment = alignment
	if border:
		for row in rows:
			for c in row:
				c.border = border 


# Spalten Namen aus Nummer erstellen 0=A
def colnum_string(t):
	string = ""
	t += 1
	while t > 0:
		t, remainder = divmod(t - 1, 26)
		string = chr(65 + remainder) + string
	return string

#Header schreiben und formatieren
def tab_write_header(ws_name, titels, row_width, X=1, frame=False):
	toal_rows = len(titels)
	if ws_name not in wb.sheetnames:
		ws = wb.create_sheet(ws_name)
	else:
		ws = wb.get_sheet_by_name(ws_name)
	font_title = Font(name='Arial' , bold=True, size=10)
	fill = PatternFill("solid", fgColor="BFBFBF")
	align_title = Alignment(horizontal='general', vertical='top')
	if frame == True:
		bd = Side(border_style='thin', color="000000")
	else:
		bd = Side(border_style=None, color="000000")
	border_def = Border(left=bd, top=bd, right=bd, bottom=bd)
	style_range(ws, 'A' + str(X) + ':' + colnum_string(toal_rows-1) + str(X) , fill=fill, font=font_title, alignment=align_title, border=border_def)
	#Spalten Titel erfassen und Breite definieren
	s = 0
	for title in titels:
		ws[colnum_string(s) + str(X)] = title
		ws.column_dimensions[colnum_string(s)].width = row_width[s]
		s = s + 1

# Daten schreiben (normal)
def tab_write_data(ws, row, X, frame=False):
	ws = wb[ws]
	font = Font(name='Arial', size=10)
	if frame == True:
		bd = Side(border_style='thin', color="000000")
	else:
		bd = Side(border_style=None, color="000000")
	border_def = Border(left=bd, top=bd, right=bd, bottom=bd)
	s = 0
	for r in row:
		ws[colnum_string(s) + str(X)] = r
		s = s + 1
	style_range(ws, 'A' + str(X) + ':' + colnum_string(s-1) + str(X), fill=None, font=font, border=border_def)
	
# Daten schreiben (Warnung)
def tab_write_data_warn(ws, row, X, frame=False):
	try:
		ws = wb[ws]
	except:
		ws = wb.create_sheet(ws) 
	font_warn = Font(name='Arial' , bold=True, size=10, color="ff0000")
	fill_warn = PatternFill("solid", fgColor="ffd7d7")
	if frame == True:
		bd = Side(border_style='thin', color="000000")
	else:
		bd = Side(border_style=None, color="000000")
	border_def = Border(left=bd, top=bd, right=bd, bottom=bd)
	s = 0
	for r in row:
		ws[colnum_string(s) + str(X)] = r
		s = s + 1
	style_range(ws, 'A' + str(X) + ':' + colnum_string(s-1) + str(X), fill=fill_warn, font=font_warn, border=border_def)
	
# Woorkbook speichern
def save_excel(out_file):
	wb.remove_sheet(wb.worksheets[0])
	wb.save(out_file)
	